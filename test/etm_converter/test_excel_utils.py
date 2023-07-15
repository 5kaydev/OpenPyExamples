import os
from typing import Any

import etm_converter.excel_utils as excel_utils
import openpyxl
import pytest


def _create_cell(value: Any):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    cell = sheet['A1']
    cell.value = value
    return cell


@pytest.mark.parametrize("cell_value, expected", [
    (None, True),
    ('', True),
    (' ', True),
    ('True', True),
    ('False', False),
    ('anything', False)
])
def test_cell_value_as_boolean(cell_value: Any, expected: str | None):
    cell = _create_cell(cell_value)
    result = excel_utils.cell_value_as_boolean(cell)
    assert result == expected


@pytest.mark.parametrize("cell_value, expected", [
    (None, None),
    ('a', 'a'),
    (1, '1')
])
def test_cell_value_as_string(cell_value: Any, expected: str | None):
    cell = _create_cell(cell_value)
    result = excel_utils.cell_value_as_string(cell)
    assert result == expected


@pytest.mark.parametrize("cell_value, expected", [
    (None, None),
    (' a ', 'a'),
    ('b', 'b'),
    (1, '1')
])
def test_stripped_cell_value(cell_value: Any, expected: str | None):
    cell = _create_cell(cell_value)
    result = excel_utils.stripped_cell_value(cell)
    assert result == expected


@pytest.mark.parametrize("input_string, expected", [
    (None, True),
    ('', True),
    ('a', False)
])
def test_is_none_or_empty(input_string: str, expected: bool):
    assert expected == excel_utils.is_none_or_empty(input_string)


def test_rewrite_excel(request, tmp_path):
    input_file = os.path.join(os.path.dirname(request.module.__file__), 'test_rewrite.xlsx')
    output_file = os.path.join(tmp_path, 'test_rewrite_out.xlsx')
    result = excel_utils.rewrite_excel(input_file, output_file)
    assert result
    workbook = openpyxl.load_workbook(output_file)
    for sheet in workbook.worksheets:
        # assert non-empty rows have been converted
        for row_index in (1, 3):
            for col_index in range(1, 6):
                if col_index != 4:
                    cell = sheet.cell(row_index, col_index)
                    assert isinstance(cell.value, str)
                    assert str(col_index) == cell.value
        # assert empty rows are left alone
        for col_index in range(1, 6):
            cell = sheet.cell(2, col_index)
            assert (cell is None) or (cell.value is None) or (cell.value == '')
    os.remove(output_file)


def test_rewrite_excel_error(request, tmp_path):
    input_file = os.path.join(os.path.dirname(request.module.__file__), 'does_not_exist.xlsx')
    output_file = os.path.join(tmp_path, 'test_rewrite_out.xlsx')
    result = excel_utils.rewrite_excel(input_file, output_file)
    assert not result

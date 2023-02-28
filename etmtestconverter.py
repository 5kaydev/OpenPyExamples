import sys

from excelutils import cell_value_as_string
import openpyxl

"""Constants for elements to parse in the Excel file"""
DO_NOT_INCLUDE = 'DONOTINCLUDE'
TESTCASE_DESCRIPTION = 2
TESTCASE_FIRST_OBJECT_NAME = 9
TESTCASE_NAME = 1


def cell_value(cell, is_input):
    value = cell_value_as_string(cell)
    if value is None:
        return '' if is_input else None
    if value == DO_NOT_INCLUDE:
        return None
    return value


def parse_json_input(filename, request_sheet):
    inputs = []
    opening_row = None
    closing_row = None
    for row_index in range(1, request_sheet.max_row + 1):
        json_element = cell_value_as_string(request_sheet.cell(row_index, 1))
        if json_element is not None:
            json_element = json_element.strip()
            if json_element == '{':
                opening_row = row_index
            if json_element == '}':
                closing_row = row_index
    if opening_row is None or closing_row is None:
        print("Missing opening or closing bracket in json request for file {0}".format(filename), file=sys.stderr)
        return None
    for column_index in range(2, request_sheet.max_column + 1):
        properties = []
        for row_index in range(opening_row + 1, closing_row):
            input_value = cell_value(request_sheet.cell(row_index, column_index), True)
            if input_value is not None:
                properties.append(cell_value_as_string(request_sheet.cell(row_index, 1))
                                  .replace(u'\xa0', '')
                                  .replace(',', '')
                                  .replace('string', input_value)
                                  .strip())
        inputs.append((cell_value_as_string(request_sheet.cell(1, column_index)),
                       '\n'.join(['"""', '{', ',\n'.join(properties), '}', '"""'])))
    return tuple(inputs)


def parse_xml_input(filename, request_sheet):
    inputs = []
    for column_index in range(3, request_sheet.max_column + 1):
        properties = []
        for row_index in range(3, request_sheet.max_row + 1):
            start_tag = cell_value_as_string(request_sheet.cell(row_index, 1))
            end_tag = cell_value_as_string(request_sheet.cell(row_index, 2))
            if end_tag is None or end_tag.strip() == '':
                properties.append(start_tag)
            else:
                input_value = cell_value(request_sheet.cell(row_index, column_index), True)
                if input_value is not None:
                    properties.append(start_tag + input_value + end_tag)
        inputs.append((cell_value_as_string(request_sheet.cell(1, column_index)),
                       '\n'.join(['"""', '\n'.join(properties), '"""'])))
    return tuple(inputs)


INPUT_PARSERS = {'Json': parse_json_input, 'XMLTagNamesStart': parse_xml_input}
INPUT_TYPES = {'Json': 'json', 'XMLTagNamesStart': 'xml'}


def parse_output(filename, validation_sheet):
    outputs = []
    for column_index in range(2, validation_sheet.max_column + 1):
        properties = []
        for row_index in range(2, validation_sheet.max_row + 1):
            expression = cell_value_as_string(validation_sheet.cell(row_index, 1))
            output_value = cell_value(validation_sheet.cell(row_index, column_index), False)
            if expression is not None and expression.strip() != '' and output_value is not None:
                properties.append((expression.strip(), output_value))
        outputs.append((cell_value_as_string(validation_sheet.cell(1, column_index)),
                        properties))
    return tuple(outputs)


def parse_workbook(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        try:
            test_data_sheet = workbook['TestData']
        except KeyError:
            print("No TestData sheet found in file {0}".format(filename), file=sys.stderr)
            return None
        testcases = (parse_testcase(filename, workbook, row)
                     for row in test_data_sheet.iter_rows(min_row=2)
                     if row[TESTCASE_DESCRIPTION].value == 'XMLWebServiceTest')
        return (testcase for testcase in testcases if testcase is not None)
    except Exception as e:
        print(e, file=sys.stderr)
        return None


def parse_testcase(filename, workbook, row):
    name = "{0}_{1}".format(cell_value_as_string(row[TESTCASE_NAME]), row[TESTCASE_NAME].row - 1)
    parameters = parse_testcase_parameters(row)
    try:
        request_sheet = workbook[parameters['RequestSheet']]
        validation_sheet = workbook[parameters['ValidationSheet']]
    except KeyError:
        print("Missing request or validation sheet found in file {0}".format(filename), file=sys.stderr)
        return None
    request_type = cell_value_as_string(request_sheet.cell(1, 1))
    if request_type not in INPUT_PARSERS.keys():
        print("Unknown request type found in file {0}, request sheet {1}".format(filename, parameters['RequestSheet']),
              file=sys.stderr)
        return None
    inputs = INPUT_PARSERS[request_type](filename, request_sheet)
    outputs = parse_output(filename, validation_sheet)
    if inputs is None or outputs is None:
        return None
    return TestCase(name, inputs, outputs, parameters, INPUT_TYPES[request_type])


def parse_testcase_parameters(row):
    return {cell_value_as_string(row[object_index]): cell_value_as_string(row[object_index + 1])
            for object_index in range(TESTCASE_FIRST_OBJECT_NAME, len(row) - 1, 2)
            if row[object_index].value is not None and row[object_index + 1].value is not None}


class TestCase:
    def __init__(self, name, inputs, outputs, parameters, request_type):
        self.name = name
        self.inputs = inputs
        self.outputs = outputs
        self.parameters = parameters
        self.request_type = request_type

    def feature(self):
        scenarios = ['Feature: {0}'.format(self.name)]
        for sc_input, sc_output in zip(self.inputs, self.outputs):
            scenarios.append(self.scenario(sc_input, sc_output))
        return '\n\n'.join(scenarios)

    def scenario(self, sc_input, sc_output):
        input_name, body = sc_input
        if self.name.startswith('S_'):
            annotation = '@SmokeTest\n'
        elif self.name.startswith('R_'):
            annotation = '@RegressionTest\n'
        else:
            annotation = ''
        lines = ['{0}Scenario: {1}'.format(annotation, input_name),
                 'Given I am a XMLWebservice client',
                 'When I send a POST request to URL "{0}" with the following {1} body'
                 .format(self.parameters['URL'], self.request_type),
                 body,
                 'And Request Header is "{0}"'.format(self.parameters['RequestHeader'])]
        output_name, outputs = sc_output
        for index in range(0, len(outputs)):
            prefix = 'Then' if index == 0 else 'And'
            expression, value = outputs[index]
            if expression.lower().strip().startswith('response code'):
                lines.append('{0} I validate that the Response Code should be {1}'
                             .format(prefix, value))
            else:
                lines.append('{0} I validate that the {1} path expression "{2}" should be "{3}"'
                             .format(prefix, self.request_type, expression, value))
        return '\n'.join(lines)

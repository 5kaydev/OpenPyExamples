import sys
from dataclasses import dataclass

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class Sheet:
    cells: list[list[str | None]]
    columns: int
    name: str
    rows: int

    def cell(self, row_index: int, column_index: int) -> str | None:
        return self.cells[row_index][column_index]


@dataclass(frozen=True)
class SpreadSheet:
    sheets: dict[str, Sheet]

    def sheet(self, sheet_name: str) -> Sheet:
        return self.sheets[sheet_name]

    def sheet_names(self) -> tuple[str]:
        return tuple(self.sheets.keys())


def _cell_value(cell: Cell) -> str | None:
    if cell is None or cell.value is None:
        return None
    return str(cell.value).strip()


def _is_cell_empty(cell: Cell) -> bool:
    value = _cell_value(cell)
    return value is None or value == '' or value.isspace()


def _is_column_empty(sheet: Worksheet, last_row: int, column_index: int) -> bool:
    for row_index in range(1, last_row + 1):
        current_cell = sheet.cell(row_index, column_index)
        if not _is_cell_empty(current_cell):
            return False
    return True


def _is_row_empty(sheet: Worksheet, row_index: int) -> bool:
    for col_index in range(1, sheet.max_column + 1):
        current_cell = sheet.cell(row_index, col_index)
        if not _is_cell_empty(current_cell):
            return False
    return True


def _last_non_empty_column(sheet: Worksheet, last_row: int) -> int:
    column_index = sheet.max_column
    while column_index > 0 and _is_column_empty(sheet, last_row, column_index):
        column_index = column_index - 1
    return column_index


def _last_non_empty_row(sheet: Worksheet) -> int:
    row_index = sheet.max_row
    if row_index == 0:
        return 0
    while row_index > 0 and _is_row_empty(sheet, row_index):
        row_index = row_index - 1
    return row_index


def load_excel(input_filename: str) -> SpreadSheet | None:
    """
    Loads an Excel file into a spreadsheet
    :param input_filename: The name of the Excel file to read.
    :return: The SpreadSheet or None in case of error
    """
    print(f'Load Excel file: {input_filename}', file=sys.stderr)
    try:
        sheets = {}
        workbook = openpyxl.load_workbook(input_filename, data_only=True)
        for sheet_name in workbook.sheetnames:
            work_sheet = workbook[sheet_name]
            print(f'Sheet: {sheet_name}, Rows: {work_sheet.max_row}, Columns: {work_sheet.max_column}', file=sys.stderr)
            rows = _last_non_empty_row(work_sheet)
            if rows > 0:
                columns = _last_non_empty_column(work_sheet, rows)
                cells = [[_cell_value(work_sheet.cell(row_index, column_index))
                          for column_index in range(1, columns + 1)]
                         for row_index in range(1, rows + 1)]
                sheets[sheet_name.strip()] = Sheet(cells, columns, sheet_name, rows)
            else:
                sheets[sheet_name.strip()] = Sheet([], 0, sheet_name, 0)
        return SpreadSheet(sheets)
    except Exception as e:
        print(e, file=sys.stderr)
        return None

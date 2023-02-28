import sys
from dataclasses import dataclass

from excelutils import stripped_cell_value
import openpyxl

SC_DATA_FIELD1 = 11
SC_OBJECT_NAME1 = 10
SC_TESTING_ACTION_FUNCTIONALITY = 6

TAF_ACTION = 'Action'
TAF_CLOSE_ALL = 'closeallbrowsers'
TAF_DATA_ENTRY = 'EnterData'
TAF_LAUNCH_AUT = 'LaunchAUT'
TAF_OBJECT_EXIST = 'ObjectExists'
TAF_VALIDATION = 'ValidateData'

TEST_ASSERTIONS = ('ObjectExists', 'ValidateData')

UI_OBJECT_BROWSER_TITLE = 15
UI_OBJECT_BROWSER_URL = 14
UI_OBJECT_CLASS_NAME = 9
UI_OBJECT_DESCRIPTIVE_PROGRAMMING = 11
UI_OBJECT_ID = 6
UI_OBJECT_INNER_TEXT = 8
UI_OBJECT_FRAME = 10
UI_OBJECT_NAME = 7
UI_OBJECT_OBJECT_NAME = 2
UI_OBJECT_PAGE_NAME = 22
UI_OBJECT_RECOVERY_SCENARIO = 13
UI_OBJECT_TAG_NAME = 4
UI_OBJECT_TIME_OUT = 12
UI_OBJECT_TYPE = 3
UI_OBJECT_XPATH = 5


def _parse_ui_object(sheet, row_index):
    """
    Parses one ui object at the given row in the given sheet
    :param sheet: The sheet containing the data
    :param row_index: The row index for the object to parse
    :return: A UIObject containing the data extracted from the row_index row of the sheet
    """

    def cell_value(column):
        return stripped_cell_value(sheet.cell(row_index, column))

    object_name = cell_value(UI_OBJECT_OBJECT_NAME)
    if object_name is None or object_name == '':
        return None
    return UIObject(
        cell_value(UI_OBJECT_BROWSER_TITLE),
        cell_value(UI_OBJECT_BROWSER_URL),
        cell_value(UI_OBJECT_CLASS_NAME),
        cell_value(UI_OBJECT_DESCRIPTIVE_PROGRAMMING),
        cell_value(UI_OBJECT_ID),
        cell_value(UI_OBJECT_INNER_TEXT),
        cell_value(UI_OBJECT_FRAME),
        cell_value(UI_OBJECT_NAME),
        object_name.lower(),
        cell_value(UI_OBJECT_PAGE_NAME),
        cell_value(UI_OBJECT_RECOVERY_SCENARIO),
        cell_value(UI_OBJECT_TAG_NAME),
        cell_value(UI_OBJECT_TIME_OUT),
        cell_value(UI_OBJECT_TYPE),
        cell_value(UI_OBJECT_XPATH)
    )


def _parse_ui_object_sheet(sheet):
    return tuple(ui_object
                 for ui_object in (_parse_ui_object(sheet, row_index)
                                   for row_index in range(2, sheet.max_row + 1))
                 if ui_object is not None)


def parse_ui_objects(filename):
    """
    Parses the given Excel file into a map of UIObjects.

    :param filename: The name of the file to parse
    :return: A map of Object names to UIObject
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        return {ui_object.object_name: ui_object
                for sheet in workbook.worksheets
                for ui_object in _parse_ui_object_sheet(sheet)
                }
    except Exception as e:
        print(e, file=sys.stderr)
        return None


def _locate_scenarios(sheet):
    """
    Identifies the various scenarios in the given sheet
    :param sheet: The sheet containing the data
    :return: a list of pairs (start, end) of row indexes for each scenario

    The end of a scenario is determined by looking at a transition from a validation action to a
    non validation action
    """
    scenarios = []
    start = 2
    validation_state = False
    for row_index in range(2, sheet.max_row + 1):
        testing_action = stripped_cell_value(sheet.cell(row_index, SC_TESTING_ACTION_FUNCTIONALITY))
        if testing_action in TEST_ASSERTIONS:
            validation_state = True
        else:
            if validation_state:
                scenarios.append((start, row_index))
                start = row_index
            validation_state = False
    if sheet.max_row > start:
        scenarios.append((start, sheet.max_row + 1))
    return scenarios


def _process_boolean_value(val):
    lower = val.lower()
    return lower if lower == 'true' or lower == 'false' else val


def _parse_action(sheet, row_index, ui_objects_map):
    action = stripped_cell_value(sheet.cell(row_index, SC_DATA_FIELD1))
    object_name = stripped_cell_value(sheet.cell(row_index, SC_OBJECT_NAME1))
    if action is None or object_name is None:
        print('Incomplete action on row {0}'.format(row_index), file=sys.stderr)
        return None
    object_name = object_name.lower()
    if object_name in ui_objects_map:
        return ActionAction(action, object_name)
    print('Object name {0} not found for action on row {1}'.format(object_name, row_index), file=sys.stderr)
    return None


def _parse_close_all_browser(sheet, row_index, ui_objects_map):
    return CloseAllBrowsersAction()


def _parse_name_value_pairs(sheet, row_index, ui_objects_map):
    inputs = []
    object_not_found = False
    i = SC_OBJECT_NAME1
    while i < sheet.max_column:
        object_name = stripped_cell_value(sheet.cell(row_index, i))
        value = stripped_cell_value(sheet.cell(row_index, i + 1))
        if object_name is not None and value is not None:
            object_name = object_name.lower()
            inputs.append((object_name, _process_boolean_value(value)))
            if object_name not in ui_objects_map:
                print('Object name {0} not found on row {1}'.format(object_name, row_index), file=sys.stderr)
                object_not_found = True
            i += 2
        else:
            break
    return () if object_not_found else tuple(inputs)


def _parse_data_entry(sheet, row_index, ui_objects_map):
    inputs = _parse_name_value_pairs(sheet, row_index, ui_objects_map)
    if inputs:
        return [DataEntryAction(object_name, value)
                for object_name, value in inputs]
    else:
        print('No data specified in EnterData action', file=sys.stderr)
        return None


def _parse_launch_aut(sheet, row_index, ui_objects_map):
    url = stripped_cell_value(sheet.cell(row_index, SC_DATA_FIELD1))
    if url is None:
        print('Missing url on LaunchAUT action on row {0}'.format(row_index), file=sys.stderr)
        return None
    else:
        return LaunchAUTAction(url)


def _parse_object_exists(sheet, row_index, ui_objects_map):
    object_name = stripped_cell_value(sheet.cell(row_index, SC_OBJECT_NAME1))
    value = stripped_cell_value(sheet.cell(row_index, SC_DATA_FIELD1))
    if object_name is None or value is None:
        print('Incomplete ObjectExists on row {0}'.format(row_index), file=sys.stderr)
        return None
    object_name = object_name.lower()
    if object_name in ui_objects_map:
        return ObjectTestAction(object_name, _process_boolean_value(value))
    print('Object name {0} not found for ObjectExists on row {1}'.format(object_name, row_index), file=sys.stderr)
    return None


def _parse_validation(sheet, row_index, ui_objects_map):
    assertions = _parse_name_value_pairs(sheet, row_index, ui_objects_map)
    if assertions:
        return [ValidationAction(object_name, value)
                for object_name, value in assertions]
    else:
        print('No assertion specified in ValidateData action', file=sys.stderr)
        return None


ACTION_PARSERS = {TAF_ACTION: _parse_action,
                  TAF_CLOSE_ALL: _parse_close_all_browser,
                  TAF_DATA_ENTRY: _parse_data_entry,
                  TAF_LAUNCH_AUT: _parse_launch_aut,
                  TAF_OBJECT_EXIST: _parse_object_exists,
                  TAF_VALIDATION: _parse_validation}


def _parse_scenario(sheet, index, row_range, ui_objects_map):
    actions = []
    start, end = row_range
    for row_index in range(start, end):
        testing_action = stripped_cell_value(sheet.cell(row_index, SC_TESTING_ACTION_FUNCTIONALITY))
        if testing_action is not None and testing_action in ACTION_PARSERS.keys():
            new_actions = ACTION_PARSERS[testing_action](sheet, row_index, ui_objects_map)
            if new_actions is None:
                return None
            if isinstance(new_actions, list):
                actions.extend(new_actions)
            else:
                actions.append(new_actions)
        else:
            print('Testing action {0} on row {1} not recognized'.format(testing_action, row_index), file=sys.stderr)
            return None
    return Scenario('_{0}'.format(index + 1), tuple(actions))


def parse_scenarios(filename, ui_objects_map):
    """
    # Parses the given Excel file into a tuple of Scenarios.
    #
    # :param filename: The name of the file to parse
    # :param ui_objects_map: The map of Object names to UIObject
    # :return: A tuple of Scenarios.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        try:
            sheet = workbook[workbook.sheetnames[0]]
        except KeyError:
            print("No sheet found in file {0}".format(filename), file=sys.stderr)
            return None
        return tuple((_parse_scenario(sheet, index, row_range, ui_objects_map)
                      for (index, row_range) in enumerate(_locate_scenarios(sheet))))
    except Exception as e:
        print(e, file=sys.stderr)
        return None


def feature(feature_name, scenarios, ui_objects_map):
    """
    Generates the content of a feature file for the given scenarios
    :param feature_name: The feature name
    :param scenarios: The scenarios in the feature
    :param ui_objects_map: The ui objects map
    :return:
    """
    sections = ['Feature: {0}'.format(feature_name)]
    for scenario in scenarios:
        sections.append(scenario.scenario(feature_name, ui_objects_map))
    return '\n\n'.join(sections)


@dataclass(frozen=True)
class UIObject:
    browser_title: str
    browser_url: str
    class_name: str
    descriptive_programming: str
    id: str
    inner_text: str
    frame: str
    name: str
    object_name: str
    page_name: str
    recovery_scenario: str
    tag_name: str
    time_out: str
    type: str
    xpath: str


@dataclass(frozen=True)
class ActionAction:
    action: str
    object_name: str

    def generate(self, ui_objects_map):
        return 'I execute the action "{0}" on object "{1}"'.format(self.action, ui_objects_map[self.object_name].xpath)


@dataclass(frozen=True)
class CloseAllBrowsersAction:
    def generate(self, ui_objects_map):
        return 'I close all browsers'


@dataclass(frozen=True)
class DataEntryAction:
    object_name: str
    value: str

    def generate(self, ui_objects_map):
        return 'I enter "{0}" in object "{1}"'.format(self.value, ui_objects_map[self.object_name].xpath)


@dataclass(frozen=True)
class LaunchAUTAction:
    url: str

    def generate(self, ui_objects_map):
        return 'I launch the application at url "{0}"'.format(self.url)


@dataclass(frozen=True)
class ObjectTestAction:
    object_name: str
    value: str

    def generate(self, ui_objects_map):
        return 'I test that object "{0}" existence is "{1}"'.format(ui_objects_map[self.object_name].xpath, self.value)


@dataclass(frozen=True)
class ValidationAction:
    object_name: str
    value: str

    def generate(self, ui_objects_map):
        return 'I validate that object "{0}" has value "{1}"'.format(ui_objects_map[self.object_name].xpath, self.value)


@dataclass(frozen=True)
class Scenario:
    name: str
    actions: tuple

    def scenario(self, feature_name, ui_objects_map):

        action_lines = [action.generate(ui_objects_map)
                        for action in self.actions]
        last_action = None
        for i in range(len(self.actions)):
            action = self.actions[i]
            if not (isinstance(action, ObjectTestAction) or isinstance(action, ValidationAction)):
                last_action = i
        lines = ['Scenario: {0}{1}'.format(feature_name, self.name),
                 '',
                 'Given {0}'.format(action_lines[0])]
        if 1 < last_action:
            lines.extend(['And {0}'.format(action_line)
                          for action_line in action_lines[1:last_action]])
        lines.append('When {0}'.format(action_lines[last_action]))
        lines.append('Then {0}'.format(action_lines[last_action + 1]))
        if last_action + 2 < len(action_lines):
            lines.extend(['And {0}'.format(action_line)
                          for action_line in action_lines[last_action + 2:]])
        return '\n'.join(lines)
